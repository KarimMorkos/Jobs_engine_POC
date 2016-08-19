import openpyxl as px
import xlrd


class ExcelReader(object):
    """
    This function return a list of coulmn inside the
    Excel workbook and append the value of each row inside
    each column to this list
    """
    @classmethod
    def readExcel(cls, fileName, sheetName, HDR=True, numOfHeaders = 1):
        wbook = px.load_workbook(fileName)
        wsheet = wbook.get_sheet_by_name(sheetName)
        if HDR != True:
            StartRow = 1
            numOfHeaders = 0
        else:
            numOfHeaders = numOfHeaders
            StartRow = numOfHeaders + 1
        result = []
        for x in range(1, (wsheet.max_column+1)):
            cols = []
            for y in range(StartRow,(wsheet.max_row+1)):
                cols.append(wsheet.cell(column=x, row=y).value)
            result.append(cols)
        return result

    @property
    def getSheetNames(self, fileName, sheetName):
        wBook = px.load_workbook(fileName)
        wSheet = wBook.get_sheet_by_name(sheetName)
        return wBook.get_sheet_names()

    @property
    def getSheetColumnsCount(self, fileName, sheetName):
        wBook = px.load_workbook(fileName)
        wSheet = wBook.get_sheet_by_name(sheetName)
        return len(wSheet.columns)

    @property
    def getSheetRowsCount(self, fileName, sheetName):
        wBook = px.load_workbook(fileName)
        wSheet = wBook.get_sheet_by_name(sheetName)
        return len(wSheet.rows)

    @classmethod
    def read_Excel(cls, filename, sheetname, HDR=True, numOfHeaders=1):
        wbook = xlrd.open_workbook(filename)
        wsheet = wbook.sheet_by_name(sheetname)
        rows = []
        for row_idx in range(1, wsheet.nrows):
            rows.append(list(wsheet.row_values(row_idx, 0, wsheet.ncols)))

        return rows
